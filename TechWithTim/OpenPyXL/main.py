from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

"""Otwórz domyślny arkusz"""
# wb = load_workbook('TwartoscRHC.xlsx')
# ws = wb.active

"""Zwróć/zmień wartośc komórki"""
# print(ws['A1'].value)
# temp = ws['A1'].value
#
# ws['A1'].value = 'Some random shit'
# print(ws['A1'].value)
#
# ws['A1'].value = temp
# print(ws['A1'].value)
#

"""Dodaj/usuń arkusz"""
# print(wb.sheetnames)
#
# wb.create_sheet('Tatus_and_Robus')
# print(wb.sheetnames)
# print(wb[wb.sheetnames[-1]])
#
# wb.remove_sheet(wb['Tatus_and_Robus'])
# print(wb.sheetnames)

# wb.save('TwartoscRHC.xlsx')

"""Utwórz arkusz i uzupełniaj wiersz po wierszu"""
# wb = Workbook()
# ws = wb.active
#
# ws.title = 'Data_sheet'
# ws.append('Mam wielkiego i niczego sobie'.split(' '))
# ws.append(['I', 'nic', 'nikomu', 'do', 'tego'])
# wb.save('Cluth.xlsx')

"""Iteruj po komórkach, printuj i podmieniaj wartości"""
# wb = load_workbook('Cluth.xlsx')
# ws = wb.active
#
# for row in range(1, 3):
#     row_value = ''
#     for col in range(1, 6):
#         char = get_column_letter(col)
#         print(ws[char + str(row)].value)
#         ws[char + str(row)].value = '* '
#         row_value += ws[char + str(row)].value
#
#     print(row_value)
#     print('__STOP__')

"""Złączanie i rozdzielanie komórek"""
# wb = load_workbook('Cluth.xlsx')
# ws = wb.active
#
# ws['A3'].value = 'Coco chanel'
# ws.merge_cells('A3:E4')
# ws.unmerge_cells('A3:E4')

"""Wstawianie i usuwanie wierszy/kolumn na danej pozycji"""
# wb = load_workbook('Cluth.xlsx')
# ws = wb.active
#
# ws.insert_rows(2)
# ws.insert_rows(4)
# ws.insert_cols(1)
#
# ws.delete_rows(2)
# ws.delete_rows(4)
# ws.delete_cols(1)

"""Przesuwanie komórek"""
# wb = load_workbook('Cluth.xlsx')
# ws = wb.active
#
# ws.move_range('A1:E2', rows=0, cols=6)
# ws.move_range('G1:K2', rows=0, cols=-6)

"""Wprowadzanie formuł i formatowanie"""
wb = load_workbook('Cluth.xlsx')
ws = wb.active

for col in range(1, 6):
    char = get_column_letter(col)
    ws[char + "3"] = f'=UPPER({char + "1"}&" "&{char + "2"})'
    ws[char + "3"].font = Font(color='FF0000', bold=True)
    ws[char + "3"].alignment = Alignment(horizontal='center')

wb.save('Cluth.xlsx')
